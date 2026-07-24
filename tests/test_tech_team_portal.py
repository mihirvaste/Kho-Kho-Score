import tempfile
import unittest
import sqlite3
from contextlib import closing
from pathlib import Path

from openpyxl import load_workbook

from backend import tech_team_app


class TechTeamPortalTests(unittest.TestCase):
    def test_load_template_expands_includes(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            template_dir = Path(tmp_dir)
            (template_dir / "partials").mkdir(parents=True, exist_ok=True)
            (template_dir / "partials" / "nav.html").write_text("NAV", encoding="utf-8")
            (template_dir / "page.html").write_text("<!-- include: partials/nav.html -->", encoding="utf-8")

            previous_dir = tech_team_app.TEMPLATE_DIR
            tech_team_app.TEMPLATE_DIR = template_dir
            try:
                rendered = tech_team_app.load_template("page.html")
            finally:
                tech_team_app.TEMPLATE_DIR = previous_dir

            self.assertIn(b"NAV", rendered)

    def test_dashboard_stats_returns_counts(self):
        tech_team_app.setup_database()
        stats = tech_team_app.get_dashboard_stats()
        self.assertIn("umpires", stats)
        self.assertIn("managers", stats)
        self.assertIn("players", stats)
        self.assertIn("tournaments", stats)

    def test_teams_template_contains_search_and_list_heading(self):
        rendered = tech_team_app.load_template("teams.html")
        self.assertIn(b"Teams", rendered)
        self.assertIn(b"Search team", rendered)

    def test_teams_template_returns_option_loader_promise_for_edit_modal(self):
        rendered = tech_team_app.load_template("teams.html")
        self.assertIn(b"return Promise.all([fetch('/api/tournaments')", rendered)
        self.assertIn(b"confirm-dialog", rendered)

    def test_manager_template_contains_confirmation_popup_styles(self):
        rendered = tech_team_app.load_template("managers.html")
        self.assertIn(b"confirm-dialog", rendered)
        self.assertIn(b"confirm-btn-yes", rendered)
        self.assertIn(b".confirm-dialog", rendered)

    def test_public_registration_templates_load_for_invite_links(self):
        for name in ["register_manager.html", "register_umpire.html", "register_player.html", "register_tournament.html"]:
            rendered = tech_team_app.load_template(name, title="Registration")
            self.assertIn(b"Registration", rendered)

    def test_tech_team_helpers_return_joined_entity_rows(self):
        tech_team_app.setup_database()
        manager_rows = tech_team_app.get_manager_rows()
        team_rows = tech_team_app.get_team_rows()
        player_rows = tech_team_app.get_player_rows()
        umpire_rows = tech_team_app.get_umpire_rows()

        self.assertTrue(manager_rows)
        self.assertIn("team_name", manager_rows[0])
        self.assertTrue(team_rows)
        self.assertIn("manager_name", team_rows[0])
        self.assertTrue(player_rows)
        self.assertIn("tournament_name", player_rows[0])
        self.assertTrue(umpire_rows)
        self.assertIn("umpire_id", umpire_rows[0])

    def test_save_draws_persists_group_and_manual_matches(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            temp_db = Path(tmp_dir) / "login.db"
            previous_db = tech_team_app.DB_PATH
            tech_team_app.DB_PATH = temp_db
            try:
                tech_team_app.setup_database()
                group_payload = {
                    "draws": [
                        {"team_a_id": "TE0001", "team_b_id": "TE0002", "group_name": "Group A", "match_number": 1, "match_status": "Scheduled", "umpire_id": "UM0001"}
                    ],
                    "scope": "group"
                }
                result = tech_team_app.save_draws("KT0001", group_payload)
                self.assertTrue(result["success"])
                with closing(sqlite3.connect(temp_db)) as db:
                    rows = db.execute("SELECT group_name, team_a_id, team_b_id FROM matches WHERE tournament_id = ?", ("KT0001",)).fetchall()
                self.assertEqual(len(rows), 1)
                self.assertEqual(rows[0][0], "Group A")

                second_group_payload = {
                    "draws": [
                        {"team_a_id": "TE0005", "team_b_id": "TE0006", "group_name": "Group B", "match_number": 1, "match_status": "Scheduled", "umpire_id": "UM0001"}
                    ],
                    "scope": "group"
                }
                result = tech_team_app.save_draws("KT0001", second_group_payload)
                self.assertTrue(result["success"])
                with closing(sqlite3.connect(temp_db)) as db:
                    rows = db.execute("SELECT team_a_id, team_b_id FROM matches WHERE tournament_id = ? AND group_name != 'Manual'", ("KT0001",)).fetchall()
                self.assertEqual(len(rows), 2)

                manual_payload = {
                    "draws": [
                        {"team_a_id": "TE0003", "team_b_id": "TE0004", "group_name": "Manual", "stage_name": "Manual", "match_number": 1, "match_status": "Scheduled", "umpire_id": "UM0002"}
                    ],
                    "scope": "manual"
                }
                result = tech_team_app.save_draws("KT0001", manual_payload)
                self.assertTrue(result["success"])
                with closing(sqlite3.connect(temp_db)) as db:
                    manual_rows = db.execute("SELECT draw_id, group_name, stage_name, team_a_id, team_b_id, umpire_id FROM matches WHERE tournament_id = ? AND group_name = 'Manual'", ("KT0001",)).fetchall()
                self.assertEqual(len(manual_rows), 1)
                self.assertEqual(manual_rows[0][1], "Manual")
                self.assertEqual(manual_rows[0][2], "Manual")
                self.assertEqual(manual_rows[0][5], "UM0002")
                self.assertTrue(manual_rows[0][0].startswith("DW"))

                second_manual_payload = {
                    "draws": [
                        {"team_a_id": "TE0005", "team_b_id": "TE0006", "group_name": "Manual", "stage_name": "Manual", "match_number": 1, "match_status": "Scheduled", "umpire_id": "UM0003"}
                    ],
                    "scope": "manual"
                }
                result = tech_team_app.save_draws("KT0001", second_manual_payload)
                self.assertTrue(result["success"])
                with closing(sqlite3.connect(temp_db)) as db:
                    manual_rows = db.execute("SELECT team_a_id, team_b_id FROM matches WHERE tournament_id = ? AND group_name = 'Manual'", ("KT0001",)).fetchall()
                self.assertEqual(len(manual_rows), 2)
            finally:
                tech_team_app.DB_PATH = previous_db

    def test_save_draws_persists_knockout_rounds(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            temp_db = Path(tmp_dir) / "login.db"
            previous_db = tech_team_app.DB_PATH
            tech_team_app.DB_PATH = temp_db
            try:
                tech_team_app.setup_database()
                result = tech_team_app.save_draws("KT0001", {
                    "scope": "knockout",
                    "draws": [
                        {"team_a_id": "TE0001", "team_b_id": "TE0002", "group_name": "Quarter Final", "stage_name": "Knockout Stage - Quarter Final", "match_number": 1, "match_status": "Scheduled"},
                        {"team_a_id": "TE0003", "team_b_id": "TE0004", "group_name": "Semi Final", "stage_name": "Knockout Stage - Semi Final", "match_number": 101, "match_status": "Scheduled"},
                    ],
                })
                self.assertTrue(result["success"])
                with closing(sqlite3.connect(temp_db)) as db:
                    rows = db.execute("SELECT stage_name, team_a_id, team_b_id FROM matches ORDER BY match_number").fetchall()
                self.assertEqual(rows, [
                    ("Knockout Stage - Quarter Final", "TE0001", "TE0002"),
                    ("Knockout Stage - Semi Final", "TE0003", "TE0004"),
                ])
            finally:
                tech_team_app.DB_PATH = previous_db

    def test_extract_match_id_from_score_sheet_route(self):
        self.assertEqual(
            tech_team_app._extract_match_id_from_path("/api/matches/TM0018/score-sheet"),
            "TM0018",
        )

    def test_score_sheet_export_populates_team_player_and_umpire_fields(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            temp_db = Path(tmp_dir) / "login.db"
            previous_db = tech_team_app.DB_PATH
            tech_team_app.DB_PATH = temp_db
            try:
                tech_team_app.setup_database()
                with closing(sqlite3.connect(temp_db)) as db:
                    db.execute("INSERT INTO teams (team_id, team_name, tournament_id, manager_id) VALUES (?, ?, ?, ?)", ("TE0001", "Alpha Team", "KT0001", None))
                    db.execute("INSERT INTO teams (team_id, team_name, tournament_id, manager_id) VALUES (?, ?, ?, ?)", ("TE0002", "Beta Team", "KT0001", None))
                    db.execute("INSERT INTO players (player_id, team_id, player_name, kkfi_number, chest_number, manager_id) VALUES (?, ?, ?, ?, ?, ?)", ("PL0001", "TE0001", "Player One", "KKFI001", 1, None))
                    db.execute("INSERT INTO players (player_id, team_id, player_name, kkfi_number, chest_number, manager_id) VALUES (?, ?, ?, ?, ?, ?)", ("PL0002", "TE0001", "Player Two", "KKFI002", 2, None))
                    db.execute("INSERT INTO players (player_id, team_id, player_name, kkfi_number, chest_number, manager_id) VALUES (?, ?, ?, ?, ?, ?)", ("PL0003", "TE0002", "Player Three", "KKFI003", 3, None))
                    db.execute("INSERT INTO players (player_id, team_id, player_name, kkfi_number, chest_number, manager_id) VALUES (?, ?, ?, ?, ?, ?)", ("PL0004", "TE0002", "Player Four", "KKFI004", 4, None))
                    db.execute("INSERT INTO umpires (umpire_id, name, kkfi_number, aadhar_photo_url, gender, age, phone, email, password_hash, blocked) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)", ("UM0001", "Umpire One", "KKFI1001", None, "M", 35, "9999999999", "umpire@example.com", "hash", 0))
                    db.execute("INSERT INTO matches (match_id, draw_id, match_number, tournament_id, group_name, stage_name, team_a_id, team_b_id, match_status, umpire_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)", ("MT0001", "DW0001", 1, "KT0001", "Group A", "Group A", "TE0001", "TE0002", "Scheduled", "UM0001"))
                    db.commit()

                output_path = tech_team_app._build_match_score_sheet(
                    "MT0001",
                    {"match_id": "MT0001", "team_a_id": "TE0001", "team_b_id": "TE0002", "stage_name": "Group A", "umpire_id": "UM0001"},
                    tech_team_app.get_team_rows(),
                    tech_team_app.get_player_rows(),
                    tech_team_app.get_umpire_rows(),
                )

                workbook = load_workbook(output_path)
                sheet = workbook[workbook.sheetnames[0]]

                self.assertEqual(sheet['C15'].value, 'TEAM: Alpha Team')
                self.assertEqual(sheet['AA15'].value, 'TEAM: Beta Team')
                self.assertEqual(sheet['C16'].value, 'Player One')
                self.assertEqual(sheet['AA16'].value, 'Player Three')
                self.assertEqual(sheet['AI50'].value, 'Umpire One')
            finally:
                tech_team_app.DB_PATH = previous_db


if __name__ == "__main__":
    unittest.main()
